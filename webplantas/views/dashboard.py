# Backend/webplantas/views/dashboard.py
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.db.models import Sum, Count, F, Q, FloatField
from django.db.models.functions import Cast, TruncDate
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal


from webplantas.models import Pedido, DetallePedido, Usuario, CatalogoPilon, RutaEntrega
from webplantas.permissions import EsPersonalViveroOAdmin


class DashboardView(APIView):
    """Vista del dashboard con estadísticas generales"""
    permission_classes = [EsPersonalViveroOAdmin]
    
    def get(self, request):
        # Estadísticas básicas
        total_pedidos = Pedido.objects.count()
        pedidos_pendientes = Pedido.objects.filter(
            estado__in=['recibido', 'confirmado', 'en_preparacion']
        ).count()
        pedidos_entregados = Pedido.objects.filter(estado='entregado').count()
        
        total_usuarios = Usuario.objects.filter(activo=True).count()
        usuarios_activos = Usuario.objects.filter(
            activo=True, 
            ultimo_acceso__gte=timezone.now() - timedelta(days=30)
        ).count()
        
        plantas_disponibles = CatalogoPilon.objects.filter(activo=True, stock__gt=0).count()
        rutas_pendientes = RutaEntrega.objects.filter(estado__in=['planificada', 'en_progreso']).count()
        
        # Ingresos del mes actual
        inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        ingresos_mes = Pedido.objects.filter(
            fecha_pedido__gte=inicio_mes,
            estado='entregado'
        ).aggregate(total=Sum('total'))['total'] or 0
        
        data = {
            'total_pedidos': total_pedidos,
            'pedidos_pendientes': pedidos_pendientes,
            'pedidos_entregados': pedidos_entregados,
            'total_usuarios': total_usuarios,
            'usuarios_activos': usuarios_activos,
            'plantas_disponibles': plantas_disponibles,
            'rutas_pendientes': rutas_pendientes,
            'ingresos_mes_actual': float(ingresos_mes)
        }
        
        return Response(data)


class DashboardEstadisticasView(APIView):
    """Vista para estadísticas completas del dashboard"""
    permission_classes = [IsAuthenticated, EsPersonalViveroOAdmin]
    
    def get(self, request):
        # 🔧 Obtener parámetros de filtro
        filtro = request.query_params.get('filtro', 'ultimos_7_dias')
        fecha_inicio_param = request.query_params.get('fecha_inicio')
        fecha_fin_param = request.query_params.get('fecha_fin')
        
        # 🔧 Usar timezone.localtime() para zona horaria correcta
        ahora_local = timezone.localtime(timezone.now())
        hoy = ahora_local.date()
        
        print(f"\n{'='*50}")
        print(f"🔍 FILTRO SELECCIONADO: {filtro}")
        print(f"📅 HOY: {hoy}")
        
        # ============= CALCULAR RANGO DE FECHAS SEGÚN FILTRO =============
        if filtro == 'personalizado' and fecha_inicio_param and fecha_fin_param:
            # Filtro personalizado
            fecha_inicio = timezone.datetime.strptime(fecha_inicio_param, '%Y-%m-%d').date()
            fecha_fin = timezone.datetime.strptime(fecha_fin_param, '%Y-%m-%d').date()
            dias_rango = (fecha_fin - fecha_inicio).days + 1
            print(f"📅 Personalizado: {fecha_inicio} a {fecha_fin} ({dias_rango} días)")
            
        elif filtro == 'ultimos_30_dias':
            fecha_inicio = hoy - timedelta(days=29)
            fecha_fin = hoy
            dias_rango = 30
            print(f"📅 Últimos 30 días: {fecha_inicio} a {fecha_fin}")
            
        elif filtro == 'este_mes':
            fecha_inicio = hoy.replace(day=1)
            fecha_fin = hoy
            dias_rango = (fecha_fin - fecha_inicio).days + 1
            print(f"📅 Este mes: {fecha_inicio} a {fecha_fin} ({dias_rango} días)")
            
        elif filtro == 'mes_pasado':
            # Primer día del mes pasado
            primer_dia_este_mes = hoy.replace(day=1)
            ultimo_dia_mes_pasado = primer_dia_este_mes - timedelta(days=1)
            fecha_inicio = ultimo_dia_mes_pasado.replace(day=1)
            fecha_fin = ultimo_dia_mes_pasado
            dias_rango = (fecha_fin - fecha_inicio).days + 1
            print(f"📅 Mes pasado: {fecha_inicio} a {fecha_fin} ({dias_rango} días)")
            
        else:  # 'ultimos_7_dias' (default)
            fecha_inicio = hoy - timedelta(days=6)
            fecha_fin = hoy
            dias_rango = 7
            print(f"📅 Últimos 7 días: {fecha_inicio} a {fecha_fin}")
        
        # ============= CREAR RANGOS DE DATETIME =============
        inicio_rango_dt = timezone.make_aware(
            timezone.datetime.combine(fecha_inicio, timezone.datetime.min.time())
        )
        fin_rango_dt = timezone.make_aware(
            timezone.datetime.combine(fecha_fin, timezone.datetime.max.time())
        )
        
        # Para las estadísticas generales (hoy, semana, mes)
        inicio_hoy = timezone.make_aware(
            timezone.datetime.combine(hoy, timezone.datetime.min.time())
        )
        fin_hoy = timezone.make_aware(
            timezone.datetime.combine(hoy, timezone.datetime.max.time())
        )
        
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        inicio_semana_dt = timezone.make_aware(
            timezone.datetime.combine(inicio_semana, timezone.datetime.min.time())
        )
        
        inicio_mes = hoy.replace(day=1)
        inicio_mes_dt = timezone.make_aware(
            timezone.datetime.combine(inicio_mes, timezone.datetime.min.time())
        )
        
        # ============= VENTAS =============
        # Ventas de hoy
        ventas_hoy_query = Pedido.objects.filter(
            fecha_pedido__gte=inicio_hoy,
            fecha_pedido__lte=fin_hoy,
            activo=True
        )
        
        ventas_hoy = ventas_hoy_query.exclude(estado='cancelado').aggregate(
            total=Sum('total')
        )['total'] or Decimal('0.00')
        
        # Ventas de la semana
        ventas_semana = Pedido.objects.filter(
            fecha_pedido__gte=inicio_semana_dt,
            activo=True
        ).exclude(estado='cancelado').aggregate(
            total=Sum('total')
        )['total'] or Decimal('0.00')
        
        # Ventas del mes
        ventas_mes = Pedido.objects.filter(
            fecha_pedido__gte=inicio_mes_dt,
            activo=True
        ).exclude(estado='cancelado').aggregate(
            total=Sum('total')
        )['total'] or Decimal('0.00')
        
        # ============= PEDIDOS =============
        pedidos_hoy = ventas_hoy_query.count()
        
        pedidos_semana = Pedido.objects.filter(
            fecha_pedido__gte=inicio_semana_dt,
            activo=True
        ).count()
        
        pedidos_mes = Pedido.objects.filter(
            fecha_pedido__gte=inicio_mes_dt,
            activo=True
        ).count()
        
        # Pedidos por estado
        pedidos_por_estado = {}
        estados = Pedido.objects.filter(activo=True).values('estado').annotate(
            total=Count('id')
        )
        for estado in estados:
            pedidos_por_estado[estado['estado']] = estado['total']
        
        # ============= TOP 5 PRODUCTOS MÁS VENDIDOS =============
        # Usar el rango del filtro seleccionado
        productos_vendidos = DetallePedido.objects.filter(
            pedido__activo=True,
            pedido__fecha_pedido__gte=inicio_rango_dt,
            pedido__fecha_pedido__lte=fin_rango_dt
        ).exclude(
            pedido__estado='cancelado'
        ).values(
            'pilon_id',
            'pilon__nombre_variedad',
            'pilon__categoria__nombre'
        ).annotate(
            cantidad_vendida=Sum('cantidad'),
            total_vendido=Sum(F('cantidad') * F('precio_unitario'))
        ).order_by('-cantidad_vendida')[:5]
        
        top_productos = [
            {
                'pilon_id': p['pilon_id'],
                'nombre_variedad': p['pilon__nombre_variedad'],
                'categoria': p['pilon__categoria__nombre'] or 'Sin categoría',
                'cantidad_vendida': p['cantidad_vendida'],
                'total_vendido': float(p['total_vendido'])
            }
            for p in productos_vendidos
        ]
        
        # ============= STOCK BAJO =============
        stock_bajo = CatalogoPilon.objects.filter(
            activo=True,
            stock__lte=F('stock_minimo')
        ).annotate(
            porcentaje_stock=Cast(F('stock'), FloatField()) / Cast(F('stock_minimo'), FloatField()) * 100
        ).values(
            'id',
            'nombre_variedad',
            'categoria__nombre',
            'stock',
            'stock_minimo',
            'porcentaje_stock'
        ).order_by('porcentaje_stock')[:10]
        
        productos_stock_bajo = [
            {
                'id': p['id'],
                'nombre_variedad': p['nombre_variedad'],
                'categoria': p['categoria__nombre'] or 'Sin categoría',
                'stock': p['stock'],
                'stock_minimo': p['stock_minimo'],
                'porcentaje_stock': round(p['porcentaje_stock'], 1)
            }
            for p in stock_bajo
        ]
        
        # ============= VENTAS DIARIAS DEL RANGO =============
        ventas_diarias = []
        
        print(f"\n📊 Generando ventas diarias para {dias_rango} días")
        
        # Generar ventas para cada día del rango
        for i in range(dias_rango):
            fecha = fecha_inicio + timedelta(days=i)
            
            # Crear rango de datetime para ese día
            inicio_dia = timezone.make_aware(
                timezone.datetime.combine(fecha, timezone.datetime.min.time())
            )
            fin_dia = timezone.make_aware(
                timezone.datetime.combine(fecha, timezone.datetime.max.time())
            )
            
            # Consultar pedidos de ese día
            ventas_dia = Pedido.objects.filter(
                fecha_pedido__gte=inicio_dia,
                fecha_pedido__lte=fin_dia,
                activo=True
            ).exclude(estado='cancelado').aggregate(
                total=Sum('total'),
                cantidad=Count('id')
            )
            
            total = float(ventas_dia['total'] or 0)
            cantidad = ventas_dia['cantidad'] or 0
            
            ventas_diarias.append({
                'fecha': fecha.strftime('%Y-%m-%d'),
                'total': total,
                'cantidad_pedidos': cantidad
            })
            
            emoji = "✅" if cantidad > 0 else "⭕"
            if cantidad > 0:
                print(f"{emoji} {fecha}: Q{total:.2f} ({cantidad} pedidos)")
        
        print(f"{'='*50}\n")
        
        # ============= RESPUESTA =============
        data = {
            'ventas': {
                'total_ventas_hoy': float(ventas_hoy),
                'total_ventas_semana': float(ventas_semana),
                'total_ventas_mes': float(ventas_mes),
                'pedidos_hoy': pedidos_hoy,
                'pedidos_semana': pedidos_semana,
                'pedidos_mes': pedidos_mes,
                'pedidos_por_estado': pedidos_por_estado,
            },
            'top_productos': top_productos,
            'stock_bajo': productos_stock_bajo,
            'ventas_diarias': ventas_diarias,
        }
        
        return Response(data)


class MetricasView(APIView):
    """Vista para métricas personalizadas por rol"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        
        if user.roles.filter(nombre_rol='Agricultor').exists():
            return self._metricas_agricultor(user)
        elif user.roles.filter(nombre_rol__in=['Administrador', 'Personal Vivero']).exists():
            return self._metricas_admin()
        else:
            return Response({'error': 'Sin permisos para ver métricas'}, status=403)
    
    def _metricas_agricultor(self, user):
        """Métricas específicas del agricultor"""
        mis_pedidos = Pedido.objects.filter(usuario=user)
        
        data = {
            'total_pedidos': mis_pedidos.count(),
            'pedidos_activos': mis_pedidos.exclude(estado__in=['entregado', 'cancelado']).count(),
            'pedidos_entregados': mis_pedidos.filter(estado='entregado').count(),
            'gasto_total': float(mis_pedidos.filter(estado='entregado').aggregate(Sum('total'))['total'] or 0),
            'notificaciones_no_leidas': user.notificaciones.filter(leida=False).count()
        }
        
        return Response(data)
    
    def _metricas_admin(self):
        """Métricas para administradores"""
        hoy = timezone.localtime(timezone.now()).date()
        
        data = {
            'pedidos_hoy': Pedido.objects.filter(fecha_pedido__date=hoy).count(),
            'rutas_hoy': RutaEntrega.objects.filter(fecha_planificada=hoy).count(),
            'entregas_hoy': Pedido.objects.filter(fecha_entrega_real=hoy).count(),
            'pedidos_pendientes_confirmacion': Pedido.objects.filter(estado='recibido').count(),
            'plantas_stock_critico': CatalogoPilon.objects.filter(
                stock__lte=F('stock_minimo'),
                activo=True
            ).count()
        }
        
        return Response(data)


class ExportarExcelView(APIView):
    """Vista para exportar estadísticas del dashboard a Excel"""
    permission_classes = [IsAuthenticated, EsPersonalViveroOAdmin]
    
    def get(self, request):
        # ✅ IMPORTS DENTRO DE LA FUNCIÓN (Evita el warning de Pylance)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from django.http import HttpResponse
        except ImportError:
            return Response({
                'error': 'openpyxl no está instalado. Ejecuta: pip install openpyxl'
            }, status=500)
        
        # Obtener parámetros de filtro (mismo que DashboardEstadisticasView)
        filtro = request.query_params.get('filtro', 'ultimos_7_dias')
        fecha_inicio_param = request.query_params.get('fecha_inicio')
        fecha_fin_param = request.query_params.get('fecha_fin')
        
        ahora_local = timezone.localtime(timezone.now())
        hoy = ahora_local.date()
        
        # Calcular rango de fechas
        if filtro == 'personalizado' and fecha_inicio_param and fecha_fin_param:
            fecha_inicio = timezone.datetime.strptime(fecha_inicio_param, '%Y-%m-%d').date()
            fecha_fin = timezone.datetime.strptime(fecha_fin_param, '%Y-%m-%d').date()
        elif filtro == 'ultimos_30_dias':
            fecha_inicio = hoy - timedelta(days=29)
            fecha_fin = hoy
        elif filtro == 'este_mes':
            fecha_inicio = hoy.replace(day=1)
            fecha_fin = hoy
        elif filtro == 'mes_pasado':
            primer_dia_este_mes = hoy.replace(day=1)
            ultimo_dia_mes_pasado = primer_dia_este_mes - timedelta(days=1)
            fecha_inicio = ultimo_dia_mes_pasado.replace(day=1)
            fecha_fin = ultimo_dia_mes_pasado
        else:  # ultimos_7_dias
            fecha_inicio = hoy - timedelta(days=6)
            fecha_fin = hoy
        
        # Crear rangos datetime
        inicio_rango_dt = timezone.make_aware(
            timezone.datetime.combine(fecha_inicio, timezone.datetime.min.time())
        )
        fin_rango_dt = timezone.make_aware(
            timezone.datetime.combine(fecha_fin, timezone.datetime.max.time())
        )
        
        # Obtener datos
        ventas_total = Pedido.objects.filter(
            fecha_pedido__gte=inicio_rango_dt,
            fecha_pedido__lte=fin_rango_dt,
            activo=True
        ).exclude(estado='cancelado').aggregate(
            total=Sum('total'),
            cantidad=Count('id')
        )
        
        # Top productos
        top_productos = DetallePedido.objects.filter(
            pedido__activo=True,
            pedido__fecha_pedido__gte=inicio_rango_dt,
            pedido__fecha_pedido__lte=fin_rango_dt
        ).exclude(
            pedido__estado='cancelado'
        ).values(
            'pilon__nombre_variedad',
            'pilon__categoria__nombre'
        ).annotate(
            cantidad_vendida=Sum('cantidad'),
            total_vendido=Sum(F('cantidad') * F('precio_unitario'))
        ).order_by('-cantidad_vendida')[:10]
        
        # Stock bajo
        stock_bajo = CatalogoPilon.objects.filter(
            activo=True,
            stock__lte=F('stock_minimo')
        ).values(
            'nombre_variedad',
            'categoria__nombre',
            'stock',
            'stock_minimo'
        ).order_by('stock')[:20]
        
        # Crear workbook
        wb = Workbook()
        
        # ========== HOJA 1: RESUMEN ==========
        ws1 = wb.active
        ws1.title = "Resumen"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws1['A1'] = "REPORTE DE ESTADÍSTICAS - VIVERO"
        ws1['A1'].font = title_font
        ws1.merge_cells('A1:C1')
        
        ws1['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        ws1['A2'].font = Font(italic=True)
        ws1.merge_cells('A2:C2')
        
        ws1['A3'] = f"Generado: {ahora_local.strftime('%d/%m/%Y %H:%M')}"
        ws1['A3'].font = Font(italic=True, size=9)
        ws1.merge_cells('A3:C3')
        
        # Resumen de ventas
        ws1['A5'] = "RESUMEN DE VENTAS"
        ws1['A5'].font = Font(bold=True, size=12)
        
        ws1.append([])  # Línea 6 vacía
        
        headers_resumen = ["Métrica", "Valor"]
        ws1.append(headers_resumen)
        
        for cell in ws1[7]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        datos_resumen = [
            ["Total de Pedidos", ventas_total['cantidad'] or 0],
            ["Total de Ventas", f"Q {float(ventas_total['total'] or 0):,.2f}"],
            ["Promedio por Pedido", f"Q {float(ventas_total['total'] or 0) / max(ventas_total['cantidad'] or 1, 1):,.2f}"],
        ]
        
        for fila in datos_resumen:
            ws1.append(fila)
        
        # Ajustar anchos
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 20
        
        # ========== HOJA 2: TOP PRODUCTOS ==========
        ws2 = wb.create_sheet("Top Productos")
        
        ws2['A1'] = "TOP 10 PRODUCTOS MÁS VENDIDOS"
        ws2['A1'].font = title_font
        ws2.merge_cells('A1:D1')
        
        ws2.append([])  # Línea 2 vacía
        
        headers_productos = ["Producto", "Categoría", "Cantidad Vendida", "Total Vendido"]
        ws2.append(headers_productos)
        
        for cell in ws2[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        for producto in top_productos:
            ws2.append([
                producto['pilon__nombre_variedad'],
                producto['pilon__categoria__nombre'] or 'Sin categoría',
                producto['cantidad_vendida'],
                f"Q {float(producto['total_vendido']):,.2f}"
            ])
        
        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 18
        
        # ========== HOJA 3: STOCK BAJO ==========
        ws3 = wb.create_sheet("Stock Bajo")
        
        ws3['A1'] = "PRODUCTOS CON STOCK BAJO"
        ws3['A1'].font = title_font
        ws3.merge_cells('A1:D1')
        
        ws3.append([])
        
        headers_stock = ["Producto", "Categoría", "Stock Actual", "Stock Mínimo"]
        ws3.append(headers_stock)
        
        for cell in ws3[3]:
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        for item in stock_bajo:
            ws3.append([
                item['nombre_variedad'],
                item['categoria__nombre'] or 'Sin categoría',
                item['stock'],
                item['stock_minimo']
            ])
        
        ws3.column_dimensions['A'].width = 35
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 15
        ws3.column_dimensions['D'].width = 15
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        nombre_archivo = f'estadisticas_vivero_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        
        wb.save(response)
        return response